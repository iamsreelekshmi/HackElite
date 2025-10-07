import streamlit as st
import pandas as pd
import json
import time
from datetime import datetime, timedelta
import base64
from io import BytesIO
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import numpy as np

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def persist_summary(summary):
    """Store latest summary in session_state for export."""
    st.session_state.last_summary = summary

def summary_to_json_bytes(summary):
    import json as _json
    return _json.dumps(summary, indent=2, default=str).encode('utf-8')

def generate_pdf_report(summary):
    """Generate a professional PDF report for the given summary."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.platypus.flowables import PageBreak
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        return None, "reportlab not installed. Install with: pip install reportlab"

    buffer = BytesIO()
    
    # Create the PDF document
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    story = []
    
    # Custom styles
    title_style = styles['Title']
    title_style.fontSize = 24
    title_style.textColor = colors.HexColor('#00478f')
    title_style.alignment = TA_CENTER
    
    heading_style = styles['Heading1']
    heading_style.fontSize = 16
    heading_style.textColor = colors.HexColor('#00478f')
    
    # Title
    story.append(Paragraph("ReleaseGate Technical Assessment Report", title_style))
    story.append(Spacer(1, 20))
    
    # Timestamp
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    story.append(Paragraph(f"Generated: {timestamp}", styles['Normal']))
    story.append(Spacer(1, 30))
    
    # Decision section
    decision = summary['decision']
    decision_color = colors.green if decision == "GO" else colors.orange if decision == "CONDITIONAL" else colors.red
    
    decision_style = styles['Heading1']
    decision_style.textColor = decision_color
    decision_style.fontSize = 20
    story.append(Paragraph(f"DECISION: {decision}", decision_style))
    story.append(Spacer(1, 20))
    
    # Risk score
    story.append(Paragraph(f"Risk Score: {summary['risk_score']}", styles['Heading2']))
    story.append(Spacer(1, 20))
    
    # KPIs section
    story.append(Paragraph("Key Performance Indicators", heading_style))
    story.append(Spacer(1, 10))
    
    kpis = summary['kpis']
    kpi_data = [['Metric', 'Value']]
    for k, v in kpis.items():
        if isinstance(v, float):
            if k.endswith('_rate') or k.endswith('_ratio'):
                kpi_data.append([k.replace('_', ' ').title(), f"{v:.2%}"])
            else:
                kpi_data.append([k.replace('_', ' ').title(), f"{v:.3f}"])
        else:
            kpi_data.append([k.replace('_', ' ').title(), str(v)])
    
    kpi_table = Table(kpi_data)
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00478f')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 30))
    
    # Issues section
    if summary['reasons']:
        story.append(Paragraph("Issues Identified", heading_style))
        story.append(Spacer(1, 10))
        for reason in summary['reasons']:
            story.append(Paragraph(f"• {reason}", styles['Normal']))
        story.append(Spacer(1, 20))
    else:
        story.append(Paragraph("No Critical Issues Detected", heading_style))
        story.append(Spacer(1, 20))
    
    # Manager comments if available
    if 'manager_comments' in st.session_state and st.session_state.manager_comments:
        story.append(Paragraph("Manager Comments", heading_style))
        story.append(Spacer(1, 10))
        story.append(Paragraph(st.session_state.manager_comments, styles['Normal']))
    
    # Build PDF
    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes, None

def send_email_report(to_email, summary, pdf_bytes=None):
    """Send email with analysis report"""
    try:
        # Email configuration (you may need to configure SMTP settings)
        smtp_server = "smtp.gmail.com"  # Configure based on your email provider
        smtp_port = 587
        from_email = "releasegate@company.com"  # Configure your sender email
        password = "your_app_password"  # Configure app password
        
        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = to_email
        msg['Subject'] = f"ReleaseGate Analysis Report - {summary['decision']}"
        
        # Email body
        body = f"""
        ReleaseGate Analysis Results
        
        Decision: {summary['decision']}
        Risk Score: {summary['risk_score']}
        Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        
        {'Issues Found:' if summary['reasons'] else 'No critical issues detected.'}
        {chr(10).join(f'• {reason}' for reason in summary['reasons']) if summary['reasons'] else ''}
        
        Best regards,
        ReleaseGate System
        """
        
        msg.attach(MIMEText(body, 'plain'))
        
        # Attach PDF if provided
        if pdf_bytes:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(pdf_bytes)
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename= releasegate_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
            )
            msg.attach(part)
        
        # Note: This is a template - actual email sending requires proper SMTP configuration
        return True, "Email would be sent successfully (SMTP configuration required)"
        
    except Exception as e:
        return False, f"Email sending failed: {str(e)}"

def get_logo_html():
    """Return HTML for logo placement"""
    return "<div class='alliance-logo'>RG</div>"

# ============================================================================
# STYLING AND CSS
# ============================================================================

def inject_custom_css():
    st.markdown("""
    <style>
    /* Main theme and colors */
    .main {
        background: #f8fafc;
        color: #1e293b;
    }
    
    .stApp {
        background: linear-gradient(135deg, #e2e8f0 0%, #f1f5f9 60%, #cbd5e1 100%);
        min-height: 100vh;
        color: #1e293b !important;
    }
    
    /* Custom logo */
    .alliance-logo {
        position: absolute;
        top: 15px;
        right: 25px;
        width: 70px;
        height: 70px;
        background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%);
        border-radius: 15px;
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 20px;
        font-weight: 700;
        color: #ffffff;
        box-shadow: 0 8px 25px rgba(15, 23, 42, 0.3);
        z-index: 1200;
        letter-spacing: 1px;
        font-family: 'Inter', sans-serif;
        border: 2px solid #3b82f6;
    }
    
    /* Header styling */
    .main-header {
        background: linear-gradient(135deg, #ffffff 0%, #f8fafc 100%);
        padding: 3rem 2.5rem;
        border-radius: 20px;
        margin: 2rem 0 2rem 0;
        box-shadow: 0 10px 30px rgba(0,0,0,0.1);
        text-align: center;
        border: 1px solid #e2e8f0;
        position: relative;
        overflow: hidden;
    }
    
    .release-gate-logo {
        font-size: 3.5rem;
        font-weight: 900;
        background: linear-gradient(45deg, #0f172a, #3b82f6, #06b6d4);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin-bottom: 0.5rem;
        letter-spacing: 2px;
        font-family: 'Inter', 'Segoe UI', sans-serif;
    }
    
    .subtitle {
        font-size: 1.4rem;
        color: #475569;
        font-weight: 600;
        margin-bottom: 0.5rem;
        letter-spacing: 0.5px;
    }
    
    /* Page navigation */
    .page-nav {
        display: flex;
        justify-content: center;
        margin: 2rem 0;
        gap: 1rem;
    }
    
    .nav-step {
        padding: 0.8rem 1.5rem;
        border-radius: 12px;
        font-weight: 600;
        border: 2px solid #e2e8f0;
        background: #f8fafc;
        color: #64748b;
        min-width: 120px;
        text-align: center;
        position: relative;
    }
    
    .nav-step.active {
        background: linear-gradient(135deg, #3b82f6 0%, #06b6d4 100%);
        color: white;
        border-color: #3b82f6;
        box-shadow: 0 4px 15px rgba(59, 130, 246, 0.3);
    }
    
    .nav-step.completed {
        background: linear-gradient(135deg, #10b981 0%, #059669 100%);
        color: white;
        border-color: #10b981;
    }
    
    /* Cards */
    .card {
        background: white;
        padding: 2rem;
        border-radius: 16px;
        box-shadow: 0 4px 20px rgba(0,0,0,0.08);
        border: 1px solid #e2e8f0;
        margin: 1.5rem 0;
    }
    
    .input-method-card {
        background: linear-gradient(135deg, #ffffff 0%, #f8fafc 100%);
        border: 2px solid #e2e8f0;
        border-radius: 16px;
        padding: 2rem;
        margin: 1rem 0;
        cursor: pointer;
        transition: all 0.3s ease;
        position: relative;
        overflow: hidden;
    }
    
    .input-method-card:hover {
        border-color: #3b82f6;
        box-shadow: 0 8px 25px rgba(59, 130, 246, 0.15);
        transform: translateY(-2px);
    }
    
    .input-method-card.selected {
        border-color: #3b82f6;
        background: linear-gradient(135deg, #eff6ff 0%, #dbeafe 100%);
        box-shadow: 0 8px 25px rgba(59, 130, 246, 0.2);
    }
    
    .method-title {
        font-size: 1.5rem;
        font-weight: 700;
        color: #1e293b;
        margin-bottom: 0.5rem;
    }
    
    .method-description {
        color: #64748b;
        font-size: 1rem;
        line-height: 1.6;
    }
    
    /* Buttons */
    .primary-button {
        background: linear-gradient(135deg, #3b82f6 0%, #06b6d4 100%);
        color: white;
        border: none;
        padding: 1rem 2rem;
        border-radius: 12px;
        font-weight: 600;
        font-size: 1.1rem;
        cursor: pointer;
        transition: all 0.3s ease;
        box-shadow: 0 4px 15px rgba(59, 130, 246, 0.3);
        width: 100%;
        margin: 1rem 0;
    }
    
    .primary-button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(59, 130, 246, 0.4);
    }
    
    /* Risk score styling */
    .risk-score-container {
        text-align: center;
        padding: 2rem;
        border-radius: 16px;
        margin: 2rem 0;
    }
    
    .risk-score-go {
        background: linear-gradient(135deg, #10b981 0%, #059669 100%);
        color: white;
        border: 3px solid #10b981;
    }
    
    .risk-score-no-go {
        background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%);
        color: white;
        border: 3px solid #ef4444;
    }
    
    .decision-text {
        font-size: 3rem;
        font-weight: 900;
        margin-bottom: 1rem;
        text-shadow: 0 2px 10px rgba(0,0,0,0.2);
    }
    
    .risk-score-text {
        font-size: 1.5rem;
        font-weight: 600;
        opacity: 0.9;
    }
    
    /* Form styling */
    .stSelectbox > div > div {
        border-radius: 12px;
        border: 2px solid #e2e8f0;
    }
    
    .stTextInput > div > div > input {
        border-radius: 12px;
        border: 2px solid #e2e8f0;
        padding: 0.8rem;
    }
    
    .stSlider > div > div > div {
        background: linear-gradient(135deg, #3b82f6 0%, #06b6d4 100%);
    }
    
    /* Manager comments */
    .manager-comments {
        background: #fffbeb;
        border: 2px solid #fbbf24;
        border-radius: 12px;
        padding: 1.5rem;
        margin: 1.5rem 0;
    }
    
    .manager-comments h4 {
        color: #92400e;
        margin-bottom: 1rem;
    }
    
    /* Success/Error messages */
    .success-message {
        background: #f0fdf4;
        border: 2px solid #22c55e;
        color: #15803d;
        padding: 1rem;
        border-radius: 12px;
        margin: 1rem 0;
    }
    
    .error-message {
        background: #fef2f2;
        border: 2px solid #ef4444;
        color: #dc2626;
        padding: 1rem;
        border-radius: 12px;
        margin: 1rem 0;
    }
    
    /* Hide Streamlit default elements */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    .stDeployButton {display:none;}
    
    </style>
    """, unsafe_allow_html=True)

# ============================================================================
# DEFAULT POLICY CONFIGURATION
# ============================================================================

DEFAULT_POLICY = {
    "pass_rate_min": 0.95,
    "high_pass_rate_min": 0.98,
    "max_blocked_tests": 0,
    "forbid_critical_or_blocked_defects": True,
    "medium_defects_max": 2,
    "minor_defects_note_threshold": 5,
    "weights": {
        "pass_rate": 0.40,
        "high_pass_rate": 0.30,
        "blocked_ratio": 0.15,
        "defect_penalty": 0.15
    }
}

# ============================================================================
# DATA PROCESSING FUNCTIONS
# ============================================================================

def calculate_defect_lifecycle_metrics(defect_data):
    """
    Calculate comprehensive defect lifecycle metrics from defect data
    """
    if len(defect_data) == 0:
        return {}
    
    total_defects = len(defect_data)
    
    # Status-based calculations
    status_counts = {}
    if 'defect_status' in defect_data.columns:
        status_counts = defect_data['defect_status'].value_counts().to_dict()
    
    # Calculate key ratios
    open_statuses = ['Open', 'In Progress', 'Reopened']
    closed_statuses = ['Resolved', 'Closed']
    rejected_status = ['Rejected']
    
    open_count = sum(status_counts.get(status, 0) for status in open_statuses)
    closed_count = sum(status_counts.get(status, 0) for status in closed_statuses)
    rejected_count = status_counts.get('Rejected', 0)
    
    # Age analysis if defect_age_days column exists
    age_metrics = {}
    if 'defect_age_days' in defect_data.columns:
        active_defects = defect_data[defect_data['defect_status'].isin(open_statuses)]
        if len(active_defects) > 0:
            age_metrics = {
                'avg_active_defect_age': float(active_defects['defect_age_days'].mean()),
                'max_active_defect_age': int(active_defects['defect_age_days'].max()),
                'min_active_defect_age': int(active_defects['defect_age_days'].min())
            }
    
    # Assignee distribution if available
    assignee_metrics = {}
    if 'defect_assignee' in defect_data.columns:
        assignee_counts = defect_data['defect_assignee'].value_counts().to_dict()
        assignee_metrics = {
            'defect_assignee_distribution': assignee_counts,
            'total_assignees': len(assignee_counts)
        }
    
    return {
        'defect_status_counts': status_counts,
        'total_defects': total_defects,
        'open_defects': open_count,
        'closed_defects': closed_count,
        'rejected_defects': rejected_count,
        'open_rate': open_count / total_defects if total_defects > 0 else 0,
        'closure_rate': closed_count / total_defects if total_defects > 0 else 0,
        'rejection_rate': rejected_count / total_defects if total_defects > 0 else 0,
        **age_metrics,
        **assignee_metrics
    }

def parse_excel_file(file):
    """Parse uploaded Excel file and extract test data."""
    try:
        # Read the Excel file
        df = pd.read_excel(file, sheet_name=0)
        
        print(f"DEBUG EXCEL: Original columns: {df.columns.tolist()}")
        print(f"DEBUG EXCEL: Data shape: {df.shape}")
        
        # Expected columns (adjust based on your Excel format)
        required_cols = ['test_id', 'status', 'priority']
        
        # Normalize column names
        df.columns = df.columns.str.lower().str.strip()
        
        print(f"DEBUG EXCEL: Normalized columns: {df.columns.tolist()}")
        
        # Basic validation
        if not all(col in df.columns for col in required_cols):
            st.error(f"Excel file must contain columns: {required_cols}")
            print(f"DEBUG EXCEL: Missing required columns!")
            return None
        
        # Process the data
        total_tests = len(df)
        passed_tests = len(df[df['status'].str.lower() == 'pass'])
        failed_tests = len(df[df['status'].str.lower() == 'fail'])
        blocked_tests = len(df[df['status'].str.lower() == 'blocked'])
        
        # High priority tests
        high_priority_df = df[df['priority'].str.lower().isin(['high', 'critical'])]
        high_total = len(high_priority_df)
        high_passed = len(high_priority_df[high_priority_df['status'].str.lower() == 'pass'])
        
        # Calculate KPIs
        kpis = {
            'total_tests': total_tests,
            'passed': passed_tests,
            'failed': failed_tests,
            'blocked': blocked_tests,
            'pass_rate': passed_tests / total_tests if total_tests > 0 else 0,
            'blocked_ratio': blocked_tests / total_tests if total_tests > 0 else 0,
            'high_total': high_total,
            'high_passed': high_passed,
            'high_pass_rate': high_passed / high_total if high_total > 0 else None
        }
        
        # Process defect data if defect_severity column exists (after normalization)
        print(f"DEBUG EXCEL: Checking for defect_severity in columns: {df.columns.tolist()}")
        print(f"DEBUG EXCEL: 'defect_severity' in df.columns: {'defect_severity' in df.columns}")
        
        if 'defect_severity' in df.columns:
            print(f"DEBUG EXCEL: ✅ defect_severity column found!")
            defect_data = df[df['defect_severity'].notna() & (df['defect_severity'] != '')]
            print(f"DEBUG EXCEL: Rows with defect data: {len(defect_data)}")
            
            # Debug: Show actual defect_severity values
            print(f"DEBUG EXCEL: Raw defect_severity values: {defect_data['defect_severity'].tolist()}")
            print(f"DEBUG EXCEL: Defect severity value counts: {defect_data['defect_severity'].value_counts().to_dict()}")
            
            defect_counts = {
                'Critical': len(defect_data[defect_data['defect_severity'].str.lower() == 'critical']),
                'Medium': len(defect_data[defect_data['defect_severity'].str.lower() == 'medium']),
                'Minor': len(defect_data[defect_data['defect_severity'].str.lower() == 'minor']),
                'Blocked': len(defect_data[defect_data['defect_severity'].str.lower() == 'blocked'])
            }
            kpis['defect_counts'] = defect_counts
            
            # Extract detailed defect information including business impact
            defect_details = []
            for idx, row in defect_data.iterrows():
                defect_info = {
                    'severity': row.get('defect_severity', ''),
                    'type': row.get('defect_type', ''),
                    'status': row.get('defect_status', ''),
                    'business_impact': row.get('business_impact', ''),
                    'impact_level': row.get('impact', ''),
                    'likelihood': row.get('likelihood', ''),
                    'detection_phase': row.get('detection_phase', ''),
                    'root_cause': row.get('root_cause', ''),
                    'defect_id': row.get('defect_id', ''),
                    'test_name': row.get('test_name', '')
                }
                defect_details.append(defect_info)
            
            kpis['defect_details'] = defect_details
            print(f"DEBUG EXCEL: Extracted {len(defect_details)} detailed defect records")
            
            # Process defect lifecycle data if defect_status column exists
            if 'defect_status' in df.columns:
                print(f"DEBUG EXCEL: ✅ defect_status column found!")
                
                # Calculate defect lifecycle metrics
                defect_lifecycle = calculate_defect_lifecycle_metrics(defect_data)
                kpis.update(defect_lifecycle)
                
                print(f"DEBUG EXCEL: Defect lifecycle metrics added: {defect_lifecycle}")
            else:
                print(f"DEBUG EXCEL: ℹ️ defect_status column not found - basic defect analysis only")
            
            print(f"DEBUG EXCEL: Defect counts added: {defect_counts}")
        else:
            print(f"DEBUG EXCEL: ❌ defect_severity column NOT found!")
        
        print(f"DEBUG EXCEL: Final KPIs keys: {list(kpis.keys())}")
        print(f"DEBUG EXCEL: Final KPIs defect_counts: {kpis.get('defect_counts', 'NOT_FOUND')}")
        return kpis
        
    except Exception as e:
        print(f"DEBUG EXCEL: ERROR in parse_excel_file: {str(e)}")
        import traceback
        traceback.print_exc()
        st.error(f"Error processing Excel file: {str(e)}")
        return None

def generate_excel_template():
    """Generate Excel template for test data upload with defect details"""
    try:
        import pandas as pd
        from io import BytesIO
        
        # Create comprehensive sample data structure with defect details
        template_data = {
            # Basic test information
            'test_id': ['TEST_0001', 'TEST_0002', 'TEST_0003', 'TEST_0004', 'TEST_0005', 'TEST_0006'],
            'test_name': ['Login Authentication Test', 'Payment Processing Test', 'User Registration Test', 'API Response Test', 'Data Validation Test', 'Security Authorization Test'],
            'status': ['pass', 'pass', 'fail', 'pass', 'fail', 'fail'],
            'priority': ['high', 'critical', 'medium', 'low', 'high', 'critical'],
            'execution_time': [45, 120, 30, 15, 60, 90],
            'environment': ['production', 'staging', 'production', 'staging', 'production', 'uat'],
            'module': ['authentication', 'payment', 'user_mgmt', 'api', 'validation', 'security'],
            'assignee': ['John Doe', 'Jane Smith', 'Bob Wilson', 'Alice Brown', 'Mike Davis', 'Sarah Lee'],
            'comments': ['All assertions passed', 'Transaction completed successfully', 'Validation error on email field', 'Response time within limits', 'Form validation failed', 'Authentication token expired'],
            
            # Enhanced defect information (for failed tests only)
            'defect_severity': ['', '', 'Medium', '', 'High', 'Critical'],
            'defect_type': ['', '', 'Functional', '', 'Data', 'Security'],
            'impact': ['', '', 'Medium', '', 'High', 'High'],
            'likelihood': ['', '', 'Likely', '', 'Very Likely', 'Possible'],
            'detection_phase': ['', '', 'System Testing', '', 'UAT', 'Production'],
            'root_cause': ['', '', 'Requirements Gap', '', 'Code Logic Error', 'Configuration Issue'],
            'defect_status': ['', '', 'Open', '', 'In Progress', 'Resolved'],
            'defect_id': ['', '', 'DEF_001', '', 'DEF_002', 'DEF_003'],
            'business_impact': ['', '', 'User cannot complete registration', '', 'Data integrity compromised', 'Security vulnerability exposed']
        }
        
        df = pd.DataFrame(template_data)
        
        # Create Excel file in memory
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Main data sheet
            df.to_excel(writer, sheet_name='Test_Results', index=False)
            
            # Enhanced instructions sheet
            instructions_data = {
                'Column Name': [
                    'test_id', 'test_name', 'status', 'priority', 'execution_time', 
                    'environment', 'module', 'assignee', 'comments',
                    'defect_severity', 'defect_type', 'impact', 'likelihood', 
                    'detection_phase', 'root_cause', 'defect_status', 'defect_id', 'business_impact'
                ],
                'Required': [
                    'Yes', 'No', 'Yes', 'Yes', 'No', 
                    'No', 'No', 'No', 'No',
                    'For Failed Tests', 'Optional', 'Optional', 'Optional', 
                    'Optional', 'Optional', 'Optional', 'Optional', 'Optional'
                ],
                'Valid Values': [
                    'Unique identifier (e.g., TEST_0001)',
                    'Descriptive test name',
                    'pass, fail, blocked, not_executed',
                    'low, medium, high, critical',
                    'Execution time in seconds (numeric)',
                    'staging, production, dev, uat',
                    'Module/component name',
                    'Test assignee name',
                    'Additional notes or comments',
                    'Critical, High, Medium, Low',
                    'Functional, Performance, Security, UI, Data, Integration',
                    'High, Medium, Low',
                    'Very Likely, Likely, Possible, Unlikely',
                    'Unit Testing, Integration Testing, System Testing, UAT, Production',
                    'Code Logic Error, Requirements Gap, Design Flaw, Configuration Issue, Third Party, Environment',
                    'Open, In Progress, Resolved, Closed, Rejected, Reopened',
                    'Unique defect identifier (e.g., DEF_001)',
                    'Description of business impact'
                ],
                'Description': [
                    'Unique test case identifier',
                    'Human-readable test case name',
                    'Test execution result status',
                    'Test case priority level',
                    'Time taken to execute the test',
                    'Environment where test was executed',
                    'Module or component being tested',
                    'Person assigned to the test',
                    'Any additional notes or failure details',
                    '🐛 Severity level of defect (leave empty for passed tests)',
                    '🐛 Type/category of defect found',
                    '🐛 Business impact level of the defect',
                    '🐛 Probability of defect occurring in production',
                    '🐛 Phase where defect was detected',
                    '🐛 Root cause category of the defect',
                    '🐛 Current status of defect resolution',
                    '🐛 Unique identifier for the defect',
                    '🐛 Detailed business impact description'
                ]
            }
            
            instructions_df = pd.DataFrame(instructions_data)
            instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
            
            # Defect analysis guide sheet
            defect_guide_data = {
                'Analysis Type': [
                    'Basic Analysis', 'Enhanced Defect Analysis', 'Risk Scoring', 'Quality Gates'
                ],
                'Required Columns': [
                    'test_id, status, priority',
                    'defect_severity (for failed tests)',
                    'defect_severity + impact + likelihood + detection_phase',
                    'All defect columns for comprehensive assessment'
                ],
                'Features Enabled': [
                    'Pass/Fail rates, Basic metrics',
                    'Defect distribution, Severity analysis',
                    'Risk calculations, Likelihood assessment',
                    'Go/No-go decisions, Risk thresholds'
                ],
                'Benefits': [
                    'Quick test overview',
                    'Defect tracking and analysis',
                    'Accurate risk assessment',
                    'Release readiness decisions'
                ]
            }
            
            defect_guide_df = pd.DataFrame(defect_guide_data)
            defect_guide_df.to_excel(writer, sheet_name='Defect_Analysis_Guide', index=False)
            
            # Get workbook and worksheet objects for formatting
            workbook = writer.book
            worksheet = writer.sheets['Test_Results']
            instructions_sheet = writer.sheets['Instructions']
            
            # Format headers
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            for col in range(1, len(instructions_df.columns) + 1):
                cell = instructions_sheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column widths
            for sheet in [worksheet, instructions_sheet]:
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
        
        buffer.seek(0)
        return buffer.getvalue()
        
    except Exception as e:
        st.error(f"Error generating Excel template: {str(e)}")
        return None

def call_genai_api(api_url, payload):
    """Call GenAI API for test data analysis."""
    try:
        import requests
        
        # Check if this is the Allianz GenAI lab endpoint
        if 'genai-lab.srv.allianz' in api_url:
            # Allianz GenAI lab specific headers and payload format
            headers = {
                'Content-Type': 'application/json',
                'User-Agent': 'ReleaseGate/1.0',
                'X-Source': 'ReleaseGate-QualityAssessment'
            }
            
            # Format payload for GenAI lab
            genai_payload = {
                "query": f"Analyze test results for project {payload.get('project_id', 'unknown')} build {payload.get('build_number', 'unknown')} in {payload.get('environment', 'unknown')} environment",
                "context": {
                    "project_id": payload.get('project_id'),
                    "build_number": payload.get('build_number'),
                    "environment": payload.get('environment'),
                    "analysis_depth": payload.get('analysis_depth', 'standard'),
                    "include_recommendations": payload.get('include_recommendations', True)
                },
                "response_format": "structured_test_data"
            }
            
            # For demo purposes, simulate GenAI lab response
            # In real implementation, you would make the actual API call
            st.info("🤖 **Demo Mode**: Simulating GenAI lab response. In production, this would fetch real data from the Allianz GenAI lab.")
            
            # Simulate realistic test data that might come from GenAI analysis
            import random
            random.seed(42)  # For consistent demo data
            
            total_tests = random.randint(80, 150)
            pass_rate = random.uniform(0.85, 0.95)
            passed = int(total_tests * pass_rate)
            failed = random.randint(3, 12)
            blocked = total_tests - passed - failed
            
            high_priority_total = random.randint(15, 30)
            high_priority_passed = int(high_priority_total * random.uniform(0.90, 0.98))
            
            # Simulate defect data
            defect_counts = {
                'Critical': random.randint(0, 2),
                'Medium': random.randint(1, 4),
                'Blocked': random.randint(0, 1),
                'Minor': random.randint(3, 8)
            }
            
            kpis = {
                'total_tests': total_tests,
                'passed': passed,
                'failed': failed,
                'blocked': blocked,
                'pass_rate': passed / total_tests,
                'blocked_ratio': blocked / total_tests,
                'high_total': high_priority_total,
                'high_passed': high_priority_passed,
                'high_pass_rate': high_priority_passed / high_priority_total,
                'defect_counts': defect_counts
            }
            
            return kpis
            
        else:
            # Generic API endpoint
            headers = {
                'Content-Type': 'application/json',
                'Authorization': 'Bearer your-api-key'  # Configure your API key
            }
            
            response = requests.post(api_url, json=payload, headers=headers, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                # Process API response and extract KPIs
                # This is a template - adjust based on your API response format
                kpis = {
                    'total_tests': data.get('total_tests', 0),
                    'passed': data.get('passed_tests', 0),
                    'failed': data.get('failed_tests', 0),
                    'blocked': data.get('blocked_tests', 0),
                    'pass_rate': data.get('pass_rate', 0),
                    'blocked_ratio': data.get('blocked_ratio', 0),
                    'high_pass_rate': data.get('high_priority_pass_rate'),
                    'defect_counts': data.get('defect_counts', {})
                }
                return kpis
            else:
                st.error(f"API call failed with status {response.status_code}")
                return None
            
    except Exception as e:
        st.error(f"Error calling GenAI API: {str(e)}")
        return None

def assess_release_readiness(kpis, policy):
    """Assess release readiness based on KPIs and policy thresholds."""
    reasons = []
    limitations = []
    risk_score = 0.0
    
    # Pass rate assessment
    if kpis["pass_rate"] < policy["pass_rate_min"]:
        reasons.append(f"Pass rate {kpis['pass_rate']:.2%} below minimum {policy['pass_rate_min']:.2%}")
        risk_score += policy["weights"]["pass_rate"]
    else:
        risk_score += policy["weights"]["pass_rate"] * (1 - kpis["pass_rate"])
    
    # High priority pass rate
    if kpis.get("high_pass_rate") is not None:
        if kpis["high_pass_rate"] < policy["high_pass_rate_min"]:
            reasons.append(f"High priority pass rate {kpis['high_pass_rate']:.2%} below minimum {policy['high_pass_rate_min']:.2%}")
            risk_score += policy["weights"]["high_pass_rate"]
        else:
            risk_score += policy["weights"]["high_pass_rate"] * (1 - kpis["high_pass_rate"])
    else:
        limitations.append("High priority pass rate not available")
    
    # Blocked tests
    if kpis["blocked"] > policy["max_blocked_tests"]:
        reasons.append(f"{kpis['blocked']} blocked tests (max allowed: {policy['max_blocked_tests']})")
        risk_score += policy["weights"]["blocked_ratio"]
    else:
        risk_score += policy["weights"]["blocked_ratio"] * kpis["blocked_ratio"]
    
    # Defect priorities assessment
    defect_penalty = 0
    if kpis.get("defect_counts") is not None:
        defect_counts = kpis["defect_counts"]
        
        # Critical and blocked defects (forbidden if policy says so)
        if policy["forbid_critical_or_blocked_defects"]:
            crit = defect_counts.get("Critical", 0)
            blocked_defects = defect_counts.get("Blocked", 0)
            if crit > 0 or blocked_defects > 0:
                reasons.append(f"{crit} critical and {blocked_defects} blocked defects found (forbidden)")
                defect_penalty += 1
        
        # Medium defects
        medium = defect_counts.get("Medium", 0)
        if medium > policy["medium_defects_max"]:
            reasons.append(f"{medium} medium defects (max allowed: {policy['medium_defects_max']})")
            defect_penalty += 0.5
        
        # Minor defects (warning threshold)
        minor = defect_counts.get("Minor", 0)
        if minor > policy["minor_defects_note_threshold"]:
            reasons.append(f"{minor} minor defects (note: threshold {policy['minor_defects_note_threshold']})")
            defect_penalty += 0.2
        
        # Add defect penalty to risk score
        risk_score += policy["weights"]["defect_penalty"] * min(defect_penalty, 1)
    else:
        limitations.append("Defect priority data missing; defect gating skipped")
    
    # Decision logic with new threshold
    if risk_score >= 0.2:
        decision = "NO-GO"
    elif reasons:
        decision = "CONDITIONAL"
    else:
        decision = "GO"
    
    return {
        "decision": decision,
        "risk_score": round(risk_score, 3),
        "reasons": reasons,
        "kpis": kpis,
        "policy": policy,
        "limitations": "; ".join(limitations) if limitations else ""
    }

# ============================================================================
# PAGE FUNCTIONS
# ============================================================================

def page_1_input_selection():
    """Page 1: Input method selection and GenAI configuration"""
    
    st.markdown("### 📋 Step 1: Configure Input Method")
    st.markdown("Select how you want to provide test data and configure the integration.")
    
    # Input method selection (3 columns)
    col1, col2, col3 = st.columns(3)
    
    with col1:
        excel_selected = st.session_state.get('input_method', '') == 'excel'
        if st.button("📊 Excel File Upload", key="excel_method", help="Upload test results from Excel file"):
            st.session_state.input_method = 'excel'
            st.rerun()
        
        if excel_selected:
            st.markdown("""
            <div class="input-method-card selected">
                <div class="method-title">📊 Excel File Upload</div>
                <div class="method-description">
                    Upload your test results in Excel format. Perfect for teams using traditional test management tools.
                </div>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown("""
            <div class="input-method-card">
                <div class="method-title">📊 Excel File Upload</div>
                <div class="method-description">
                    Upload your test results in Excel format. Perfect for teams using traditional test management tools.
                </div>
            </div>
            """, unsafe_allow_html=True)
    
    with col2:
        api_selected = st.session_state.get('input_method', '') == 'api'
        if st.button("🔗 GenAI API Integration", key="api_method", help="Connect to GenAI lab for real-time data"):
            st.session_state.input_method = 'api'
            st.rerun()
        
        if api_selected:
            st.markdown("""
            <div class="input-method-card selected">
                <div class="method-title">🔗 GenAI API Integration</div>
                <div class="method-description">
                    Connect directly to your GenAI lab for real-time test data analysis.
                </div>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown("""
            <div class="input-method-card">
                <div class="method-title">🔗 GenAI API Integration</div>
                <div class="method-description">
                    Connect directly to your GenAI lab for real-time test data analysis.
                </div>
            </div>
            """, unsafe_allow_html=True)
    
    with col3:
        manual_selected = st.session_state.get('input_method', '') == 'manual'
        if st.button("✍️ Manual Data Entry", key="manual_method", help="Enter test data manually"):
            st.session_state.input_method = 'manual'
            st.rerun()
        
        if manual_selected:
            st.markdown("""
            <div class="input-method-card selected">
                <div class="method-title">✍️ Manual Data Entry</div>
                <div class="method-description">
                    Enter test metrics manually for quick assessments or small-scale testing.
                </div>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown("""
            <div class="input-method-card">
                <div class="method-title">✍️ Manual Data Entry</div>
                <div class="method-description">
                    Enter test metrics manually for quick assessments or small-scale testing.
                </div>
            </div>
            """, unsafe_allow_html=True)
    
    # GenAI lab link input (only show for API method)
    if st.session_state.get('input_method') == 'api':
        st.markdown("### 🔗 GenAI Lab Configuration")
        genai_url = st.text_input(
            "GenAI Lab API Endpoint",
            placeholder="https://genai-lab.srv.allianz/app/agents/chat/agent/...",
            help="Enter the complete URL to your GenAI lab API endpoint",
            value=st.session_state.get('genai_url', 'https://genai-lab.srv.allianz/app/agents/chat/agent/7cf3cec9-afe6-4276-8ffa-be76fa65ba74')
        )
        st.session_state.genai_url = genai_url
        
        # Show helpful information about the GenAI lab integration
        st.info("🤖 **GenAI Lab Integration**: This endpoint connects to Allianz's GenAI lab for intelligent test data analysis and insights.")
    
    # Continue button
    st.markdown("---")
    if st.session_state.get('input_method'):
        if st.session_state.get('input_method') == 'api' and not st.session_state.get('genai_url'):
            st.warning("⚠️ Please provide the GenAI lab API endpoint to continue.")
        else:
            if st.button("Continue to Configuration ➡️", key="continue_to_config"):
                st.session_state.current_page = 2
                st.rerun()
    else:
        st.info("👆 Please select an input method to continue.")

def page_2_configuration():
    """Page 2: Configuration and threshold settings"""
    
    st.markdown("### ⚙️ Step 2: Configure Analysis Parameters")
    st.markdown("Set your quality thresholds and upload data based on your selected input method.")
    
    # Show configuration based on input method
    input_method = st.session_state.get('input_method', '')
    
    if input_method == 'excel':
        # Excel file upload and threshold configuration
        col1, col2 = st.columns([1, 1])
        
        with col1:
            st.markdown("#### 📁 Upload Test Results")
            
            # Excel template download
            st.markdown("**📋 Need a template with defect analysis support?**")
            st.info("💡 **Sample Template Test Summary**: Includes defect severity, impact, likelihood, detection phase, and other fields for comprehensive defect analysis!")
            template_bytes = generate_excel_template()
            if template_bytes:
                st.download_button(
                    label="📥 Download Sample Template Test Summary",
                    data=template_bytes,
                    file_name="sample_template_test_summary.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="Download sample template with test summary and defect analysis fields for quality assessment"
                )
            
            st.markdown("---")
            
            uploaded_file = st.file_uploader(
                "Choose Excel file",
                type=['xlsx', 'xls'],
                help="Upload Excel file with test results"
            )
            
            if uploaded_file:
                with st.spinner("Processing Excel file..."):
                    kpis = parse_excel_file(uploaded_file)
                    if kpis:
                        st.session_state.kpis = kpis
                        st.success("✅ Excel file processed successfully!")
                        
                        # Show preview of data
                        st.markdown("**Data Preview:**")
                        preview_data = {
                            "Total Tests": kpis['total_tests'],
                            "Passed": kpis['passed'],
                            "Failed": kpis['failed'],
                            "Blocked": kpis['blocked'],
                            "Pass Rate": f"{kpis['pass_rate']:.2%}"
                        }
                        st.json(preview_data)
        
        with col2:
            st.markdown("#### 🎯 Quality Thresholds")
            
            # Basic threshold settings
            pass_rate_min = st.slider(
                "Minimum Pass Rate (%)",
                min_value=0,
                max_value=100,
                value=int(st.session_state.get('pass_rate_min', DEFAULT_POLICY['pass_rate_min']) * 100),
                help="Overall test pass rate threshold"
            ) / 100
            
            high_pass_rate_min = st.slider(
                "High Priority Pass Rate (%)",
                min_value=0,
                max_value=100,
                value=int(st.session_state.get('high_pass_rate_min', DEFAULT_POLICY['high_pass_rate_min']) * 100),
                help="Critical test cases pass rate"
            ) / 100
            
            max_blocked_tests = st.number_input(
                "Max Blocked Tests",
                min_value=0,
                value=st.session_state.get('max_blocked_tests', DEFAULT_POLICY['max_blocked_tests']),
                step=1,
                help="Maximum acceptable blocked test cases"
            )
            
            # Defect management settings
            st.markdown("**🐛 Defect Management:**")
            forbid_critical = st.checkbox(
                "Forbid Critical/Blocked Defects", 
                value=st.session_state.get('forbid_critical', DEFAULT_POLICY['forbid_critical_or_blocked_defects']),
                help="Automatically block release for critical or blocked defects"
            )
            
            medium_defects_max = st.number_input(
                "Max Medium Defects",
                min_value=0,
                value=st.session_state.get('medium_defects_max', DEFAULT_POLICY['medium_defects_max']),
                step=1,
                help="Maximum allowed medium priority defects"
            )
            
            minor_defects_threshold = st.number_input(
                "Minor Defects Warning Threshold",
                min_value=0,
                value=st.session_state.get('minor_defects_threshold', DEFAULT_POLICY['minor_defects_note_threshold']),
                step=1,
                help="Threshold for minor defects warning"
            )
            
            # Store comprehensive policy
            policy = DEFAULT_POLICY.copy()
            policy.update({
                "pass_rate_min": pass_rate_min,
                "high_pass_rate_min": high_pass_rate_min,
                "max_blocked_tests": int(max_blocked_tests),
                "forbid_critical_or_blocked_defects": forbid_critical,
                "medium_defects_max": int(medium_defects_max),
                "minor_defects_note_threshold": int(minor_defects_threshold)
            })
            
            st.session_state.policy = policy
    
    elif input_method == 'api':
        # API configuration with full interactive settings
        col1, col2 = st.columns([1, 1])
        
        with col1:
            st.markdown("#### 🔗 API Configuration")
            genai_url = st.session_state.get('genai_url', '')
            st.text_input("GenAI Lab URL", value=genai_url, disabled=True)
            
            # Show connection status
            if genai_url:
                if 'genai-lab.srv.allianz' in genai_url:
                    st.success("🔗 Connected to Allianz GenAI Lab")
                else:
                    st.info("🔗 Custom GenAI endpoint configured")
            
            # API parameters
            st.markdown("**API Parameters:**")
            project_id = st.text_input("Project ID", placeholder="Enter project identifier (e.g., RELEASE_GATE_2024)")
            build_number = st.text_input("Build Number", placeholder="Enter build number (e.g., 1.2.3-SNAPSHOT)")
            environment = st.selectbox("Environment", ["staging", "production", "dev", "uat", "integration"])
            
            # Additional GenAI parameters
            st.markdown("**GenAI Analysis Settings:**")
            analysis_depth = st.selectbox("Analysis Depth", ["standard", "detailed", "comprehensive"])
            include_recommendations = st.checkbox("Include AI Recommendations", value=True)
            
            if st.button("🔄 Fetch Data from GenAI Lab", type="primary"):
                if project_id and build_number:
                    with st.spinner("Fetching data from GenAI lab..."):
                        payload = {
                            "project_id": project_id,
                            "build_number": build_number,
                            "environment": environment,
                            "analysis_depth": analysis_depth,
                            "include_recommendations": include_recommendations
                        }
                        kpis = call_genai_api(genai_url, payload)
                        if kpis:
                            st.session_state.kpis = kpis
                            st.success("✅ Data fetched successfully from GenAI lab!")
                            
                            # Show GenAI insights if available
                            if include_recommendations:
                                st.markdown("**🤖 GenAI Insights:**")
                                st.info("AI analysis suggests focusing on high-priority test stabilization and defect triage for optimal release readiness.")
                        else:
                            st.error("❌ Failed to fetch data from GenAI lab")
                else:
                    st.warning("⚠️ Please provide Project ID and Build Number")
        
        with col2:
            st.markdown("#### 🎯 Quality Thresholds")
            
            # Basic threshold configuration
            pass_rate_min = st.slider(
                "Minimum Pass Rate (%)",
                min_value=0,
                max_value=100,
                value=int(st.session_state.get('pass_rate_min', DEFAULT_POLICY['pass_rate_min']) * 100)
            ) / 100
            
            high_pass_rate_min = st.slider(
                "High Priority Pass Rate (%)",
                min_value=0,
                max_value=100,
                value=int(st.session_state.get('high_pass_rate_min', DEFAULT_POLICY['high_pass_rate_min']) * 100)
            ) / 100
            
            max_blocked_tests = st.number_input(
                "Max Blocked Tests",
                min_value=0,
                value=st.session_state.get('max_blocked_tests', DEFAULT_POLICY['max_blocked_tests']),
                step=1
            )
            
            # Defect management settings
            st.markdown("**🐛 Defect Management:**")
            forbid_critical = st.checkbox(
                "Forbid Critical/Blocked Defects", 
                value=st.session_state.get('forbid_critical', DEFAULT_POLICY['forbid_critical_or_blocked_defects']),
                help="Automatically block release for critical or blocked defects"
            )
            
            medium_defects_max = st.number_input(
                "Max Medium Defects",
                min_value=0,
                value=st.session_state.get('medium_defects_max', DEFAULT_POLICY['medium_defects_max']),
                step=1,
                help="Maximum allowed medium priority defects"
            )
            
            minor_defects_threshold = st.number_input(
                "Minor Defects Warning Threshold",
                min_value=0,
                value=st.session_state.get('minor_defects_threshold', DEFAULT_POLICY['minor_defects_note_threshold']),
                step=1,
                help="Threshold for minor defects warning"
            )
            
            # Advanced interactive settings for API method
            st.markdown("#### 🔧 Advanced Settings")
            
            risk_weights = st.expander("Risk Weight Configuration")
            with risk_weights:
                st.markdown("**Configure how different factors contribute to the overall risk score:**")
                col1_w, col2_w = st.columns(2)
                with col1_w:
                    pass_rate_weight = st.slider(
                        "Pass Rate Weight", 
                        0.0, 1.0, 
                        st.session_state.get('pass_rate_weight', DEFAULT_POLICY['weights']['pass_rate']), 
                        0.05,
                        help="Weight for overall pass rate in risk calculation"
                    )
                    high_pass_weight = st.slider(
                        "High Priority Weight", 
                        0.0, 1.0, 
                        st.session_state.get('high_pass_weight', DEFAULT_POLICY['weights']['high_pass_rate']), 
                        0.05,
                        help="Weight for high priority tests in risk calculation"
                    )
                with col2_w:
                    blocked_weight = st.slider(
                        "Blocked Tests Weight", 
                        0.0, 1.0, 
                        st.session_state.get('blocked_weight', DEFAULT_POLICY['weights']['blocked_ratio']), 
                        0.05,
                        help="Weight for blocked tests in risk calculation"
                    )
                    defect_weight = st.slider(
                        "Defect Weight", 
                        0.0, 1.0, 
                        st.session_state.get('defect_weight', DEFAULT_POLICY['weights']['defect_penalty']), 
                        0.05,
                        help="Weight for defects in risk calculation"
                    )
                
                # Normalize weights to sum to 1
                total_weight = pass_rate_weight + high_pass_weight + blocked_weight + defect_weight
                if total_weight > 0:
                    pass_rate_weight /= total_weight
                    high_pass_weight /= total_weight
                    blocked_weight /= total_weight
                    defect_weight /= total_weight
                    st.info(f"Weights normalized to sum to 1.0 (Total was: {total_weight:.2f})")
            
            # Store comprehensive configuration
            policy = DEFAULT_POLICY.copy()
            policy.update({
                "pass_rate_min": pass_rate_min,
                "high_pass_rate_min": high_pass_rate_min,
                "max_blocked_tests": int(max_blocked_tests),
                "forbid_critical_or_blocked_defects": forbid_critical,
                "medium_defects_max": int(medium_defects_max),
                "minor_defects_note_threshold": int(minor_defects_threshold),
                "weights": {
                    "pass_rate": pass_rate_weight,
                    "high_pass_rate": high_pass_weight,
                    "blocked_ratio": blocked_weight,
                    "defect_penalty": defect_weight
                }
            })
            
            st.session_state.policy = policy
    
    elif input_method == 'manual':
        # Manual data entry
        col1, col2 = st.columns([1, 1])
        
        with col1:
            st.markdown("#### ✍️ Manual Test Data Entry")
            
            # Display Quality Thresholds prominently
            st.info("🎯 **Quality Thresholds for GO Decision:**\n"
                   "• Pass Rate: ≥ 85%\n"
                   "• Blocked Tests: ≤ 5%\n" 
                   "• Critical Defects: 0\n"
                   "• Current defaults set for **Risk Score 0** (Perfect GO)")
            
            # Initialize session state for manual entry values (Perfect GO defaults for Risk Score 0)
            if 'manual_total_tests' not in st.session_state:
                st.session_state.manual_total_tests = 15
            if 'manual_passed_tests' not in st.session_state:
                st.session_state.manual_passed_tests = 15  # 100% pass rate for risk score 0
            if 'manual_failed_tests' not in st.session_state:
                st.session_state.manual_failed_tests = 0   # 0 failed tests for perfect score
            if 'manual_blocked_tests' not in st.session_state:
                st.session_state.manual_blocked_tests = 0
            
            # Basic test metrics with proper bounds checking
            total_tests = st.number_input(
                "Total Tests", 
                min_value=1, 
                value=st.session_state.manual_total_tests, 
                step=1,
                key="total_tests_input"
            )
            st.session_state.manual_total_tests = total_tests
            
            # Passed tests
            max_passed = total_tests
            default_passed = min(st.session_state.manual_passed_tests, max_passed)
            passed_tests = st.number_input(
                "Passed Tests", 
                min_value=0, 
                value=default_passed, 
                step=1, 
                max_value=max_passed,
                key="passed_tests_input"
            )
            st.session_state.manual_passed_tests = passed_tests
            
            # Failed tests
            max_failed = total_tests - passed_tests
            default_failed = min(st.session_state.manual_failed_tests, max_failed)
            failed_tests = st.number_input(
                "Failed Tests", 
                min_value=0, 
                value=default_failed, 
                step=1, 
                max_value=max_failed,
                key="failed_tests_input"
            )
            st.session_state.manual_failed_tests = failed_tests
            
            # Blocked tests
            max_blocked = total_tests - passed_tests - failed_tests
            default_blocked = min(st.session_state.manual_blocked_tests, max_blocked)
            blocked_tests = st.number_input(
                "Blocked Tests", 
                min_value=0, 
                value=default_blocked, 
                step=1, 
                max_value=max_blocked,
                key="blocked_tests_input"
            )
            st.session_state.manual_blocked_tests = blocked_tests
            
            # High priority tests
            st.markdown("**High Priority Tests:**")
            
            # Initialize session state for high priority values (GO Release Candidate defaults)
            if 'manual_high_total' not in st.session_state:
                st.session_state.manual_high_total = 7
            if 'manual_high_passed' not in st.session_state:
                st.session_state.manual_high_passed = 6
            
            max_high_total = total_tests
            default_high_total = min(st.session_state.manual_high_total, max_high_total)
            high_total = st.number_input(
                "High Priority Total", 
                min_value=0, 
                value=default_high_total, 
                step=1, 
                max_value=max_high_total,
                key="high_total_input"
            )
            st.session_state.manual_high_total = high_total
            
            max_high_passed = high_total
            default_high_passed = min(st.session_state.manual_high_passed, max_high_passed)
            high_passed = st.number_input(
                "High Priority Passed", 
                min_value=0, 
                value=default_high_passed, 
                step=1, 
                max_value=max_high_passed,
                key="high_passed_input"
            )
            st.session_state.manual_high_passed = high_passed
            
            # Calculate and store KPIs
            if st.button("📊 Calculate KPIs", type="primary"):
                # Validate inputs
                if passed_tests + failed_tests + blocked_tests != total_tests:
                    st.error("❌ Error: Passed + Failed + Blocked tests must equal Total tests")
                elif high_passed > high_total:
                    st.error("❌ Error: High priority passed cannot exceed high priority total")
                else:
                    # Build defect counts using session state values
                    defect_counts = {
                        'Critical': st.session_state.get('manual_critical_defects', 0),
                        'Medium': st.session_state.get('manual_medium_defects', 0),
                        'Minor': st.session_state.get('manual_minor_defects', 0),
                        'Blocked': st.session_state.get('manual_blocked_defects', 0)
                    }
                    
                    # Calculate total defects from session state
                    total_critical = st.session_state.get('manual_critical_defects', 0)
                    total_medium = st.session_state.get('manual_medium_defects', 0)
                    total_minor = st.session_state.get('manual_minor_defects', 0)
                    total_blocked_defects = st.session_state.get('manual_blocked_defects', 0)
                    total_entered_defects = total_critical + total_medium + total_minor + total_blocked_defects
                    
                    # Get defect lifecycle values from session state
                    open_defects = st.session_state.get('manual_open_defects', 0)
                    closed_defects = st.session_state.get('manual_closed_defects', 0)
                    rejected_defects = st.session_state.get('manual_rejected_defects', 0)
                    
                    kpis = {
                        'total_tests': total_tests,
                        'passed': passed_tests,
                        'failed': failed_tests,
                        'blocked': blocked_tests,
                        'pass_rate': passed_tests / total_tests if total_tests > 0 else 0,
                        'blocked_ratio': blocked_tests / total_tests if total_tests > 0 else 0,
                        'high_total': high_total,
                        'high_passed': high_passed,
                        'high_pass_rate': high_passed / high_total if high_total > 0 else None,
                        'defect_counts': defect_counts,
                        # Defect lifecycle metrics
                        'open_defects': open_defects,
                        'closed_defects': closed_defects,
                        'rejected_defects': rejected_defects,
                        'open_rate': open_defects / total_entered_defects if total_entered_defects > 0 else 0,
                        'closure_rate': closed_defects / total_entered_defects if total_entered_defects > 0 else 0,
                        'rejection_rate': rejected_defects / total_entered_defects if total_entered_defects > 0 else 0,
                        'avg_active_defect_age': 0,  # Default value for now
                        'total_assignees': 0  # Default value for now
                    }
                    st.session_state.kpis = kpis
                    st.success("✅ KPIs calculated successfully!")
                    
                    # Show calculated metrics
                    st.markdown("**Calculated Metrics:**")
                    st.write(f"- Pass Rate: {kpis['pass_rate']:.2%}")
                    st.write(f"- Blocked Ratio: {kpis['blocked_ratio']:.2%}")
                    if kpis['high_pass_rate'] is not None:
                        st.write(f"- High Priority Pass Rate: {kpis['high_pass_rate']:.2%}")
                    if total_entered_defects > 0:
                        st.write(f"- Total Defects: {total_entered_defects}")
                        st.write(f"- Defect Open Rate: {kpis['open_rate']:.1%}")
        
        with col2:
            st.markdown("#### 🐛 Defect Data Entry")
            
            # Display Defect Quality Thresholds
            st.success("✅ **Defect Quality Thresholds:**\n"
                      "• Critical: Must be 0\n"
                      "• Medium: ≤ 5 acceptable\n"
                      "• Current defaults = **Perfect Quality** (0 defects)")
            
            st.markdown("**Defect Counts by Severity:**")
            
            # Initialize session state for defect values (Perfect GO defaults for Risk Score 0)
            if 'manual_critical_defects' not in st.session_state:
                st.session_state.manual_critical_defects = 0
            if 'manual_medium_defects' not in st.session_state:
                st.session_state.manual_medium_defects = 0    # 0 medium defects for perfect score
            if 'manual_minor_defects' not in st.session_state:
                st.session_state.manual_minor_defects = 0
            if 'manual_blocked_defects' not in st.session_state:
                st.session_state.manual_blocked_defects = 0
            
            critical_defects = st.number_input(
                "Critical Defects", 
                min_value=0, 
                value=st.session_state.manual_critical_defects, 
                step=1,
                key="critical_defects_input"
            )
            st.session_state.manual_critical_defects = critical_defects
            
            medium_defects = st.number_input(
                "Medium Defects", 
                min_value=0, 
                value=st.session_state.manual_medium_defects, 
                step=1,
                key="medium_defects_input"
            )
            st.session_state.manual_medium_defects = medium_defects
            
            minor_defects = st.number_input(
                "Minor Defects", 
                min_value=0, 
                value=st.session_state.manual_minor_defects, 
                step=1,
                key="minor_defects_input"
            )
            st.session_state.manual_minor_defects = minor_defects
            
            blocked_defects = st.number_input(
                "Blocked Defects", 
                min_value=0, 
                value=st.session_state.manual_blocked_defects, 
                step=1,
                key="blocked_defects_input"
            )
            st.session_state.manual_blocked_defects = blocked_defects
            
            st.markdown("**Defect Lifecycle Status:**")
            
            # Initialize session state for defect lifecycle values (Perfect GO defaults for Risk Score 0)
            if 'manual_open_defects' not in st.session_state:
                st.session_state.manual_open_defects = 0
            if 'manual_closed_defects' not in st.session_state:
                st.session_state.manual_closed_defects = 0    # 0 total defects for perfect score
            if 'manual_rejected_defects' not in st.session_state:
                st.session_state.manual_rejected_defects = 0
            
            total_entered_defects = critical_defects + medium_defects + minor_defects + blocked_defects
            
            open_defects = st.number_input(
                "Open/In Progress Defects", 
                min_value=0, 
                value=min(st.session_state.manual_open_defects, total_entered_defects), 
                step=1,
                max_value=total_entered_defects,
                key="open_defects_input",
                help="Defects currently open or in progress"
            )
            st.session_state.manual_open_defects = open_defects
            
            max_closed = total_entered_defects - open_defects
            closed_defects = st.number_input(
                "Closed/Resolved Defects", 
                min_value=0, 
                value=min(st.session_state.manual_closed_defects, max_closed), 
                step=1,
                max_value=max_closed,
                key="closed_defects_input",
                help="Defects that have been resolved and closed"
            )
            st.session_state.manual_closed_defects = closed_defects
            
            max_rejected = total_entered_defects - open_defects - closed_defects
            rejected_defects = st.number_input(
                "Rejected Defects", 
                min_value=0, 
                value=min(st.session_state.manual_rejected_defects, max_rejected), 
                step=1,
                max_value=max_rejected,
                key="rejected_defects_input",
                help="Defects rejected as invalid or not a defect"
            )
            st.session_state.manual_rejected_defects = rejected_defects
            
            # Additional defect metrics
            st.markdown("**Additional Defect Metrics:**")
            
            avg_defect_age = st.number_input(
                "Average Age of Open Defects (days)", 
                min_value=0, 
                value=0, 
                step=1,
                key="avg_defect_age_input",
                help="Average number of days open defects have been active"
            )
            
            total_assignees = st.number_input(
                "Number of Team Members with Assigned Defects", 
                min_value=0, 
                value=0, 
                step=1,
                key="total_assignees_input",
                help="How many team members currently have defects assigned"
            )
            
            # Validation and summary
            remaining_defects = total_entered_defects - open_defects - closed_defects - rejected_defects
            if remaining_defects != 0:
                st.warning(f"⚠️ Defect status counts don't match total defects. Missing: {remaining_defects}")
            
            if total_entered_defects > 0:
                st.markdown("**📊 Defect Metrics Summary:**")
                col_a, col_b, col_c = st.columns(3)
                with col_a:
                    st.metric("Open Rate", f"{(open_defects/total_entered_defects)*100:.1f}%")
                with col_b:
                    st.metric("Closure Rate", f"{(closed_defects/total_entered_defects)*100:.1f}%")
                with col_c:
                    st.metric("Rejection Rate", f"{(rejected_defects/total_entered_defects)*100:.1f}%")
            
            # Basic threshold settings
            pass_rate_min = st.slider(
                "Minimum Pass Rate (%)",
                min_value=0,
                max_value=100,
                value=int(st.session_state.get('pass_rate_min', DEFAULT_POLICY['pass_rate_min']) * 100),
                help="Overall test pass rate threshold"
            ) / 100
            
            high_pass_rate_min = st.slider(
                "High Priority Pass Rate (%)",
                min_value=0,
                max_value=100,
                value=int(st.session_state.get('high_pass_rate_min', DEFAULT_POLICY['high_pass_rate_min']) * 100),
                help="Critical test cases pass rate"
            ) / 100
            
            max_blocked_tests = st.number_input(
                "Max Blocked Tests",
                min_value=0,
                value=st.session_state.get('max_blocked_tests', DEFAULT_POLICY['max_blocked_tests']),
                step=1,
                help="Maximum acceptable blocked test cases"
            )
            
            # Defect management settings
            st.markdown("**🐛 Defect Management:**")
            forbid_critical = st.checkbox(
                "Forbid Critical/Blocked Defects", 
                value=st.session_state.get('forbid_critical', DEFAULT_POLICY['forbid_critical_or_blocked_defects']),
                help="Automatically block release for critical or blocked defects"
            )
            
            medium_defects_max = st.number_input(
                "Max Medium Defects",
                min_value=0,
                value=st.session_state.get('medium_defects_max', DEFAULT_POLICY['medium_defects_max']),
                step=1,
                help="Maximum allowed medium priority defects"
            )
            
            minor_defects_threshold = st.number_input(
                "Minor Defects Warning Threshold",
                min_value=0,
                value=st.session_state.get('minor_defects_threshold', DEFAULT_POLICY['minor_defects_note_threshold']),
                step=1,
                help="Threshold for minor defects warning"
            )
            
            # Defect data entry (optional)
            st.markdown("**🐛 Defect Data (Optional):**")
            col_def1, col_def2 = st.columns(2)
            with col_def1:
                critical_defects = st.number_input("Critical Defects", min_value=0, value=0, step=1)
                medium_defects = st.number_input("Medium Defects", min_value=0, value=0, step=1)
            with col_def2:
                blocked_defects = st.number_input("Blocked Defects", min_value=0, value=0, step=1)
                minor_defects = st.number_input("Minor Defects", min_value=0, value=0, step=1)
            
            # Store comprehensive policy
            policy = DEFAULT_POLICY.copy()
            policy.update({
                "pass_rate_min": pass_rate_min,
                "high_pass_rate_min": high_pass_rate_min,
                "max_blocked_tests": int(max_blocked_tests),
                "forbid_critical_or_blocked_defects": forbid_critical,
                "medium_defects_max": int(medium_defects_max),
                "minor_defects_note_threshold": int(minor_defects_threshold)
            })
            
            st.session_state.policy = policy
            
            # Store defect data if entered
            if critical_defects + medium_defects + blocked_defects + minor_defects > 0:
                if 'kpis' in st.session_state:
                    st.session_state.kpis['defect_counts'] = {
                        'Critical': critical_defects,
                        'Medium': medium_defects,
                        'Blocked': blocked_defects,
                        'Minor': minor_defects
                    }
    
    # Navigation buttons
    st.markdown("---")
    col1, col2 = st.columns([1, 1])
    
    with col1:
        if st.button("⬅️ Back to Input Selection", key="back_to_input"):
            st.session_state.current_page = 1
            st.rerun()
    
    with col2:
        # Check if we have the required data to proceed
        has_data = 'kpis' in st.session_state and 'policy' in st.session_state
        if has_data:
            if st.button("Continue to Analysis ➡️", key="continue_to_analysis", type="primary"):
                st.session_state.current_page = 3
                st.rerun()
        else:
            st.button("Continue to Analysis ➡️", disabled=True, help="Please configure and load data first")

def page_3_analysis_results():
    """Page 3: Analysis results with all capabilities"""
    
    st.markdown("### 📊 Step 3: Analysis Results")
    
    # Check if we have the required data
    if 'kpis' not in st.session_state or 'policy' not in st.session_state:
        st.error("❌ No data available. Please go back and configure your input.")
        if st.button("⬅️ Back to Configuration"):
            st.session_state.current_page = 2
            st.rerun()
        return
    
    # Perform analysis
    kpis = st.session_state.kpis
    policy = st.session_state.policy
    summary = assess_release_readiness(kpis, policy)
    persist_summary(summary)
    
    # Main decision display
    decision = summary['decision']
    risk_score = summary['risk_score']
    
    # Decision styling based on risk score and decision
    if risk_score >= 0.2:
        decision_class = "risk-score-no-go"
        decision_text = "🚫 NO-GO"
        decision_emoji = "🚫"
    else:
        decision_class = "risk-score-go"
        decision_text = "✅ GO"
        decision_emoji = "✅"
    
    st.markdown(f"""
    <div class="{decision_class} risk-score-container">
        <div class="decision-text">{decision_emoji} {decision}</div>
        <div class="risk-score-text">Risk Score: {risk_score:.3f}</div>
        <div style="margin-top: 1rem; font-size: 1.1rem;">
            {"Release is blocked due to high risk" if risk_score >= 0.2 else "Release is approved with acceptable risk"}
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Executive Summary - Brief Results Overview
    st.markdown("### 📋 Executive Summary")
    
    # Build brief summary text
    total_tests = kpis['total_tests']
    pass_rate = kpis['pass_rate']
    critical_issues = len(summary['reasons']) if summary['reasons'] else 0
    
    # Defect summary
    defect_summary = ""
    if kpis.get('defect_counts'):
        defects = kpis['defect_counts']
        total_defects = sum(defects.values())
        if total_defects > 0:
            critical_defects = defects.get('Critical', 0)
            blocked_defects = defects.get('Blocked', 0)
            high_severity = critical_defects + blocked_defects
            defect_summary = f"**Defects:** {total_defects} total ({high_severity} high-severity)" if high_severity > 0 else f"**Defects:** {total_defects} total (low-medium severity)"
        else:
            defect_summary = "**Defects:** None identified"
    
    # Quality assessment
    quality_status = "🟢 HIGH" if pass_rate >= 0.95 else "🟡 MEDIUM" if pass_rate >= 0.85 else "🔴 LOW"
    
    summary_text = f"""
    **🎯 Test Results:** {kpis['passed']}/{total_tests} tests passed ({pass_rate:.1%} success rate)  
    **📊 Quality Level:** {quality_status}  
    {defect_summary}  
    **⚠️ Critical Issues:** {critical_issues} blocking concerns identified  
    **🏁 Recommendation:** {'Proceed with release' if decision == 'GO' else 'Release requires attention - address issues before deployment'}
    """
    
    st.markdown(summary_text)
    
    # Quick action items if there are issues
    if summary['reasons']:
        with st.expander("🔧 Action Items Required", expanded=False):
            for i, reason in enumerate(summary['reasons'], 1):
                st.markdown(f"**{i}.** {reason}")
    else:
        st.success("✅ **All quality gates passed** - Release is ready for deployment")
    
    # Risk Score Graph
    st.markdown("### 📈 Risk Score Visualization")
    fig = go.Figure()
    
    # Create gauge chart
    fig.add_trace(go.Indicator(
        mode = "gauge+number+delta",
        value = risk_score,
        domain = {'x': [0, 1], 'y': [0, 1]},
        title = {'text': "Risk Score"},
        delta = {'reference': 0.2},
        gauge = {
            'axis': {'range': [None, 1]},
            'bar': {'color': "darkblue"},
            'steps': [
                {'range': [0, 0.2], 'color': "lightgreen"},
                {'range': [0.2, 1], 'color': "lightcoral"}
            ],
            'threshold': {
                'line': {'color': "red", 'width': 4},
                'thickness': 0.75,
                'value': 0.2
            }
        }
    ))
    
    fig.update_layout(height=400)
    st.plotly_chart(fig, use_container_width=True)
    
    # Additional Visualizations
    st.markdown("### 📊 Detailed Analytics Dashboard")
    
    # Create multiple visualizations
    viz_col1, viz_col2 = st.columns(2)
    
    with viz_col1:
        # Test Status Distribution (Pie Chart)
        st.markdown("#### 🎯 Test Status Distribution")
        status_data = {
            'Status': ['Passed', 'Failed', 'Blocked'],
            'Count': [kpis['passed'], kpis['failed'], kpis['blocked']],
            'Percentage': [
                kpis['passed']/kpis['total_tests']*100,
                kpis['failed']/kpis['total_tests']*100,
                kpis['blocked']/kpis['total_tests']*100
            ]
        }
        
        fig_pie = px.pie(
            values=status_data['Count'],
            names=status_data['Status'],
            title="Test Execution Results",
            color_discrete_map={'Passed': '#10b981', 'Failed': '#ef4444', 'Blocked': '#f59e0b'}
        )
        fig_pie.update_layout(height=350)
        st.plotly_chart(fig_pie, use_container_width=True)
        
        # High Priority vs Overall Comparison
        if kpis.get('high_pass_rate') is not None:
            st.markdown("#### 📈 Priority Comparison")
            comparison_data = pd.DataFrame({
                'Category': ['Overall Tests', 'High Priority Tests'],
                'Pass Rate': [kpis['pass_rate'] * 100, kpis['high_pass_rate'] * 100],
                'Total Tests': [kpis['total_tests'], kpis.get('high_total', 0)]
            })
            
            fig_bar = px.bar(
                comparison_data,
                x='Category',
                y='Pass Rate',
                title="Pass Rate Comparison",
                color='Category',
                color_discrete_map={'Overall Tests': '#3b82f6', 'High Priority Tests': '#ef4444'},
                text='Pass Rate'
            )
            fig_bar.update_traces(texttemplate='%{text:.1f}%', textposition='outside')
            fig_bar.update_layout(
                height=350, 
                yaxis_title="Pass Rate (%)", 
                showlegend=False,
                yaxis=dict(range=[0, 105])
            )
            st.plotly_chart(fig_bar, use_container_width=True)
    
    with viz_col2:
        # Risk Score Breakdown
        st.markdown("#### ⚖️ Risk Score Breakdown")
        
        # Calculate individual risk components
        risk_components = {
            'Pass Rate Impact': policy["weights"]["pass_rate"] * (1 - kpis["pass_rate"]) if kpis["pass_rate"] >= policy["pass_rate_min"] else policy["weights"]["pass_rate"],
            'High Priority Impact': 0,
            'Blocked Tests Impact': policy["weights"]["blocked_ratio"] * kpis["blocked_ratio"] if kpis["blocked"] <= policy["max_blocked_tests"] else policy["weights"]["blocked_ratio"],
            'Defect Impact': 0
        }
        
        # High priority impact
        if kpis.get("high_pass_rate") is not None:
            if kpis["high_pass_rate"] >= policy["high_pass_rate_min"]:
                risk_components['High Priority Impact'] = policy["weights"]["high_pass_rate"] * (1 - kpis["high_pass_rate"])
            else:
                risk_components['High Priority Impact'] = policy["weights"]["high_pass_rate"]
        
        # Defect impact (if available)
        if kpis.get("defect_counts"):
            defect_penalty = 0
            defects = kpis["defect_counts"]
            if policy["forbid_critical_or_blocked_defects"]:
                if defects.get("Critical", 0) > 0 or defects.get("Blocked", 0) > 0:
                    defect_penalty += 1
            if defects.get("Medium", 0) > policy["medium_defects_max"]:
                defect_penalty += 0.5
            if defects.get("Minor", 0) > policy["minor_defects_note_threshold"]:
                defect_penalty += 0.2
            risk_components['Defect Impact'] = policy["weights"]["defect_penalty"] * min(defect_penalty, 1)
        
        risk_df = pd.DataFrame(list(risk_components.items()), columns=['Component', 'Risk Score'])
        risk_df = risk_df[risk_df['Risk Score'] > 0]  # Only show components with risk
        
        fig_risk = px.bar(
            risk_df,
            x='Component',
            y='Risk Score',
            title="Risk Score Components",
            color='Risk Score',
            color_continuous_scale='Reds'
        )
        fig_risk.update_layout(height=350, xaxis_tickangle=-45)
        st.plotly_chart(fig_risk, use_container_width=True)
        
        # Defect Analysis (if available)
        if kpis.get("defect_counts") and sum(kpis["defect_counts"].values()) > 0:
            st.markdown("#### 🐛 Defect Analysis")
            defect_df = pd.DataFrame(list(kpis["defect_counts"].items()), columns=['Priority', 'Count'])
            defect_df = defect_df[defect_df['Count'] > 0]  # Only show priorities with defects
            
            fig_defects = px.bar(
                defect_df,
                x='Priority',
                y='Count',
                title="Defects by Priority",
                color='Priority',
                color_discrete_map={
                    'Critical': '#dc2626',
                    'Blocked': '#7c2d12',
                    'Medium': '#f59e0b',
                    'Minor': '#10b981'
                }
            )
            fig_defects.update_layout(height=350)
            st.plotly_chart(fig_defects, use_container_width=True)
        else:
            st.info("📝 No defect data available for detailed analysis")
    
    # === ENHANCED DEFECT ANALYTICS DASHBOARD ===
    # Import and display comprehensive defect analytics if defect data is available
    if kpis.get("defect_counts") and sum(kpis["defect_counts"].values()) > 0:
        try:
            from defect_analytics import create_defect_analytics_dashboard
            create_defect_analytics_dashboard(kpis, policy)
        except ImportError:
            st.warning("⚠️ Defect analytics module not available")
        except Exception as e:
            st.error(f"❌ Error loading defect analytics: {str(e)}")
    
    # === KEY INSIGHTS AND RECOMMENDATIONS ===
    st.markdown("### 💡 Key Insights & Recommendations")
    
    # Generate insights based on analysis
    insights = []
    
    if decision == "NO-GO":
        insights.append("🚫 **Release Status**: This release is not recommended due to high risk factors.")
    elif decision == "CONDITIONAL":
        insights.append("⚠️ **Release Status**: This release has some concerns that should be addressed.")
    else:
        insights.append("✅ **Release Status**: This release meets quality criteria and is ready to go.")
    
    if summary['reasons']:
        insights.append(f"🔍 **Main Concerns**: {', '.join(summary['reasons'])}")
    
    # Display insights
    for insight in insights:
        if "🚫" in insight:
            st.error(insight)
        elif "⚠️" in insight:
            st.warning(insight)
        else:
            st.success(insight)
    
    # Quality Trends Simulation (if we had historical data)
    st.markdown("#### 📈 Quality Trend Analysis")
    st.info("💡 **Future Enhancement**: This section will show historical trends when integrated with your CI/CD pipeline")
    
    # Simulate trend data for demonstration
    import numpy as np
    from datetime import datetime, timedelta
    
    dates = [datetime.now() - timedelta(days=x) for x in range(14, 0, -1)]
    
    # Simulate some realistic trend data
    np.random.seed(42)  # For consistent demo data
    base_pass_rate = kpis['pass_rate']
    pass_rates = [max(0.7, min(0.99, base_pass_rate + np.random.normal(0, 0.03))) for _ in dates]
    risk_scores = [max(0.05, min(0.4, risk_score + np.random.normal(0, 0.05))) for _ in dates]
    
    trend_fig = make_subplots(
        rows=2, cols=1,
        subplot_titles=('Pass Rate Trend (14 Days)', 'Risk Score Trend (14 Days)'),
        vertical_spacing=0.1
    )
    
    # Pass rate trend
    trend_fig.add_trace(
        go.Scatter(
            x=dates,
            y=[pr * 100 for pr in pass_rates],
            mode='lines+markers',
            name='Pass Rate',
            line=dict(color='#10b981', width=3)
        ),
        row=1, col=1
    )
    
    # Risk score trend
    trend_fig.add_trace(
        go.Scatter(
            x=dates,
            y=risk_scores,
            mode='lines+markers',
            name='Risk Score',
            line=dict(color='#ef4444', width=3),
            fill='tonexty'
        ),
        row=2, col=1
    )
    
    # Add threshold lines
    trend_fig.add_hline(y=95, line_dash="dash", line_color="orange", row=1, col=1, annotation_text="Min Threshold")
    trend_fig.add_hline(y=0.2, line_dash="dash", line_color="red", row=2, col=1, annotation_text="No-Go Threshold")
    
    trend_fig.update_layout(
        height=500,
        showlegend=False,
        title_text="Historical Quality Trends (Simulated)"
    )
    
    trend_fig.update_yaxes(title_text="Pass Rate (%)", row=1, col=1)
    trend_fig.update_yaxes(title_text="Risk Score", row=2, col=1)
    trend_fig.update_xaxes(title_text="Date", row=2, col=1)
    
    st.plotly_chart(trend_fig, use_container_width=True)
    
    # Detailed Analysis
    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.markdown("#### 📋 Key Performance Indicators")
        kpi_data = []
        kpi_data.append({"Metric": "Total Tests", "Value": str(kpis['total_tests'])})
        kpi_data.append({"Metric": "Passed Tests", "Value": str(kpis['passed'])})
        kpi_data.append({"Metric": "Failed Tests", "Value": str(kpis['failed'])})
        kpi_data.append({"Metric": "Blocked Tests", "Value": str(kpis['blocked'])})
        kpi_data.append({"Metric": "Pass Rate", "Value": f"{kpis['pass_rate']:.2%}"})
        kpi_data.append({"Metric": "Blocked Ratio", "Value": f"{kpis['blocked_ratio']:.2%}"})
        
        if kpis.get('high_pass_rate') is not None:
            kpi_data.append({"Metric": "High Priority Pass Rate", "Value": f"{kpis['high_pass_rate']:.2%}"})
        
        kpi_df = pd.DataFrame(kpi_data)
        st.dataframe(kpi_df, use_container_width=True, hide_index=True)
    
    with col2:
        st.markdown("#### ⚠️ Issues & Recommendations")
        if summary['reasons']:
            for i, reason in enumerate(summary['reasons'], 1):
                st.markdown(f"**{i}.** {reason}")
        else:
            st.success("🎉 No critical issues detected!")
        
        if summary.get('limitations'):
            st.info(f"ℹ️ **Note:** {summary['limitations']}")
    
    # Manager Comments Section
    st.markdown("### 💬 Manager Comments")
    manager_comments = st.text_area(
        "Add your comments and observations:",
        value=st.session_state.get('manager_comments', ''),
        height=120,
        placeholder="Enter any additional context, concerns, or approval notes..."
    )
    st.session_state.manager_comments = manager_comments
    
    # Action Buttons Row
    st.markdown("### 🎯 Actions")
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        # PDF Download
        if st.button("📄 Download PDF Report", type="primary"):
            with st.spinner("Generating PDF report..."):
                pdf_bytes, error = generate_pdf_report(summary)
                if pdf_bytes:
                    st.download_button(
                        label="📥 Download PDF",
                        data=pdf_bytes,
                        file_name=f"releasegate_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                        mime="application/pdf",
                        key="download_pdf"
                    )
                else:
                    st.error(f"PDF generation failed: {error}")
    
    with col2:
        # JSON Download
        json_data = summary_to_json_bytes(summary)
        st.download_button(
            label="📋 Download JSON",
            data=json_data,
            file_name=f"releasegate_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            mime="application/json",
            key="download_json"
        )
    
    with col3:
        # Email functionality
        email_to = st.text_input("Email to:", placeholder="manager@company.com", key="email_input")
        if st.button("📧 Send Email", disabled=not email_to):
            with st.spinner("Sending email..."):
                pdf_bytes, _ = generate_pdf_report(summary)
                success, message = send_email_report(email_to, summary, pdf_bytes)
                if success:
                    st.success(f"✅ {message}")
                else:
                    st.error(f"❌ {message}")
    
    with col4:
        # Save Results
        if st.button("💾 Save Results"):
            # Save to session state or database
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            save_data = {
                'timestamp': timestamp,
                'summary': summary,
                'manager_comments': manager_comments,
                'input_method': st.session_state.get('input_method'),
                'risk_score': risk_score,
                'decision': decision
            }
            
            # Here you could implement database saving
            st.session_state.saved_results = save_data
            st.success("✅ Results saved successfully!")
    
    # Navigation
    st.markdown("---")
    col1, col2 = st.columns([1, 1])
    
    with col1:
        if st.button("⬅️ Back to Configuration", key="back_to_config"):
            st.session_state.current_page = 2
            st.rerun()
    
    with col2:
        if st.button("🔄 New Analysis", key="new_analysis"):
            # Reset session state for new analysis
            keys_to_reset = ['input_method', 'genai_url', 'kpis', 'policy', 'manager_comments']
            for key in keys_to_reset:
                if key in st.session_state:
                    del st.session_state[key]
            st.session_state.current_page = 1
            st.rerun()

# ============================================================================
# MAIN APPLICATION
# ============================================================================

def main():
    """Main application function"""
    
    # Page configuration
    st.set_page_config(
        page_title="ReleaseGate Enterprise Platform",
        page_icon="🚀",
        layout="wide",
        initial_sidebar_state="collapsed"
    )
    
    # Apply custom CSS
    inject_custom_css()
    
    # Initialize session state
    if 'current_page' not in st.session_state:
        st.session_state.current_page = 1
    
    # Main header
    st.markdown(get_logo_html() + """
    <div class="main-header">
        <div class="release-gate-logo">🚀 ReleaseGate Enterprise</div>
        <div class="subtitle">Intelligent Quality Decision Platform</div>
        <p style="color: #64748b; margin: 0; font-size: 1.1rem;">
            AI-powered release decisions through comprehensive quality analysis
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # Page navigation indicator
    current_page = st.session_state.current_page
    
    nav_step_1 = "nav-step completed" if current_page > 1 else "nav-step active" if current_page == 1 else "nav-step"
    nav_step_2 = "nav-step completed" if current_page > 2 else "nav-step active" if current_page == 2 else "nav-step"
    nav_step_3 = "nav-step active" if current_page == 3 else "nav-step"
    
    st.markdown(f"""
    <div class="page-nav">
        <div class="{nav_step_1}">
            1. Input Selection
        </div>
        <div class="{nav_step_2}">
            2. Configuration
        </div>
        <div class="{nav_step_3}">
            3. Analysis Results
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Render appropriate page
    if current_page == 1:
        page_1_input_selection()
    elif current_page == 2:
        page_2_configuration()
    elif current_page == 3:
        page_3_analysis_results()

if __name__ == "__main__":
    main()